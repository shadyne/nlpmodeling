import os
import logging
from datetime import datetime
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import seaborn as sns

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

logger = logging.getLogger(__name__)

H_FILL  = PatternFill("solid", fgColor="1F4E79")
ALT_FILL= PatternFill("solid", fgColor="EBF3FB")
W_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
N_FONT  = Font(name="Arial", size=10)
CEN     = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN    = Side(style="thin", color="B8CCE4")
BRD     = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def sh(cell, fill=None):
    """Style header cell."""
    cell.font      = W_FONT
    cell.fill      = fill or H_FILL
    cell.alignment = CEN
    cell.border    = BRD


def sc(cell, alt=False, fmt=None):
    """Style content cell."""
    cell.font      = N_FONT
    cell.alignment = CEN
    cell.border    = BRD
    if alt:
        cell.fill = ALT_FILL
    if fmt:
        cell.number_format = fmt



def make_charts(df: pd.DataFrame, chart_dir: str) -> dict:
    os.makedirs(chart_dir, exist_ok=True)
    paths = {}
    TARGET_YEARS = sorted(df["Year"].unique())
    COLORS = sns.color_palette("Blues_d", max(len(TARGET_YEARS), 3))
    sns.set_theme(style="whitegrid", font="Arial")

    fig, axes = plt.subplots(1, 3, figsize=(15, 5.5))
    fig.suptitle(
        "Mean KAM Repetition Metrics by Year",
        fontsize=14, fontweight="bold", color="#1F4E79"
    )
    for ax, (col, title) in zip(axes, [
        ("%RepKAM_prior_years", "%RepKAM_prior_years"),
        ("%RepKAM_last_year",   "%RepKAM_last_year"),
        ("KAM_Similarity",      "KAM_Similarity\n(Avg Cosine Score)"),
    ]):
        means = [df[df["Year"] == yr][col].mean() for yr in TARGET_YEARS]
        bars  = ax.bar(
            [str(y) for y in TARGET_YEARS], means,
            color=COLORS[:len(TARGET_YEARS)], width=0.55, edgecolor="white"
        )
        ax.set_title(title, fontweight="bold", fontsize=10)
        ax.set_ylim(0, 1.15)
        ax.yaxis.set_major_formatter(
            plt.FuncFormatter(lambda v, _: f"{v:.0%}")
        )
        for b in bars:
            h = b.get_height()
            if pd.notna(h) and h > 0:
                ax.text(
                    b.get_x() + b.get_width() / 2, h + 0.02,
                    f"{h:.1%}", ha="center", fontsize=9, fontweight="bold"
                )
        ax.spines[["top", "right"]].set_visible(False)
    plt.tight_layout()
    p = f"{chart_dir}/chart1_mean.png"
    plt.savefig(p, dpi=150, bbox_inches="tight")
    plt.close()
    paths["Mean Repetition Metrics"] = p

    fig, ax = plt.subplots(figsize=(10, 5.5))
    fig.suptitle(
        "Trend of KAM Metrics Across Years",
        fontsize=14, fontweight="bold", color="#1F4E79"
    )
    styles = [
        ("%RepKAM_prior_years", "%RepKAM_prior_years", "#1F4E79", "o-"),
        ("%RepKAM_last_year",   "%RepKAM_last_year",   "#2E75B6", "s--"),
        ("KAM_Similarity",      "KAM_Similarity (Avg)", "#E8A02A", "^:"),
    ]
    for col, label, color, style in styles:
        means = [df[df["Year"] == yr][col].mean() for yr in TARGET_YEARS]
        ax.plot(
            TARGET_YEARS, means, style, label=label,
            color=color, linewidth=2.5, markersize=9
        )
        for yr, m in zip(TARGET_YEARS, means):
            if pd.notna(m):
                ax.text(yr, m + 0.015, f"{m:.1%}", ha="center",
                        fontsize=9, color=color)
    ax.set_xticks(TARGET_YEARS)
    ax.set_ylim(0, 1.05)
    ax.yaxis.set_major_formatter(
        plt.FuncFormatter(lambda v, _: f"{v:.0%}")
    )
    ax.legend(fontsize=10)
    ax.spines[["top", "right"]].set_visible(False)
    plt.tight_layout()
    p = f"{chart_dir}/chart2_trend.png"
    plt.savefig(p, dpi=150, bbox_inches="tight")
    plt.close()
    paths["Trend Analysis"] = p

    fig, axes = plt.subplots(1, 2, figsize=(14, 5.5))
    fig.suptitle(
        "Distribution of KAM Similarity Scores",
        fontsize=14, fontweight="bold", color="#1F4E79"
    )
    for ax, (col, title) in zip(axes, [
        ("%RepKAM_last_year", "%RepKAM_last_year"),
        ("KAM_Similarity",    "KAM_Similarity (Avg Cosine Score)"),
    ]):
        has_plot = False
        for yr, c in zip(TARGET_YEARS, COLORS[:len(TARGET_YEARS)]):
            subset = df[(df["Year"] == yr) & df[col].notna()][col]
            if len(subset) > 1 and subset.nunique() > 1:
                try:
                    subset.plot.kde(ax=ax, label=str(yr),
                                    color=c, linewidth=2.5)
                    has_plot = True
                except Exception:
                    pass
        ax.set_title(title, fontweight="bold", fontsize=10)
        ax.set_xlim(-0.05, 1.05)
        ax.xaxis.set_major_formatter(
            plt.FuncFormatter(lambda v, _: f"{v:.0%}")
        )
        if has_plot:
            ax.legend(title="Year")
        ax.spines[["top", "right"]].set_visible(False)
    plt.tight_layout()
    p = f"{chart_dir}/chart3_dist.png"
    plt.savefig(p, dpi=150, bbox_inches="tight")
    plt.close()
    paths["Distributions"] = p

    fig, ax = plt.subplots(figsize=(10, 5.5))
    fig.suptitle(
        "Average New KAMs per Year",
        fontsize=14, fontweight="bold", color="#1F4E79"
    )
    x  = np.arange(len(TARGET_YEARS))
    w  = 0.3
    mp = [df[df["Year"] == yr]["#NewKAM_prior_years"].mean()
          for yr in TARGET_YEARS]
    ml = [df[df["Year"] == yr]["#NewKAM_last_year"].mean()
          for yr in TARGET_YEARS]
    mp = [0 if pd.isna(v) else v for v in mp]
    ml = [0 if pd.isna(v) else v for v in ml]

    b1 = ax.bar(x - w/2, mp, w, label="#NewKAM_prior_years",
                color="#1F4E79", edgecolor="white")
    b2 = ax.bar(x + w/2, ml, w, label="#NewKAM_last_year",
                color="#9DC3E6", edgecolor="white")
    ax.set_xticks(x)
    ax.set_xticklabels([str(y) for y in TARGET_YEARS])
    ax.set_ylabel("Avg # New KAMs")
    ax.legend()
    ax.spines[["top", "right"]].set_visible(False)
    for bars in [b1, b2]:
        for b in bars:
            h = b.get_height()
            if h > 0:
                ax.text(
                    b.get_x() + b.get_width() / 2, h + 0.02,
                    f"{h:.2f}", ha="center", fontsize=9
                )
    plt.tight_layout()
    p = f"{chart_dir}/chart4_new.png"
    plt.savefig(p, dpi=150, bbox_inches="tight")
    plt.close()
    paths["New KAMs"] = p

    return paths



def export_full_excel(metrics: list, output_path: str):

    if not metrics:
        print("[!] Tidak ada metrik untuk diekspor.")
        return

    df = pd.DataFrame(metrics)
    TARGET_YEARS = sorted(df["Year"].unique())
    chart_dir  = os.path.join(os.path.dirname(output_path), "charts")
    chart_paths = make_charts(df, chart_dir)

    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "KAM Analysis Data"
    HEADERS = [
        "Company ID", "Year", "#KAMs",
        "%RepKAM_prior_years", "%RepKAM_last_year",
        "KAM_Similarity\n(Avg Cosine vs t-1)",
        "#NewKAM_prior_years", "#NewKAM_last_year",
    ]

    ws.append(["KAM Disclosure Similarity Analysis"])
    ws.merge_cells("A1:H1")
    ws["A1"].font      = Font(name="Arial", bold=True, size=14, color="1F4E79")
    ws["A1"].alignment = CEN
    ws.row_dimensions[1].height = 28

    ws.append([
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')} "
        f"| Method: TF-IDF Cosine Similarity (bigram, sublinear_tf) "
        f"| Repetition threshold: 0.70"
    ])
    ws.merge_cells("A2:H2")
    ws["A2"].font      = Font(name="Arial", size=9, italic=True, color="666666")
    ws["A2"].alignment = CEN

    ws.append([])
    ws.append(HEADERS)
    for c, h in enumerate(HEADERS, 1):
        sh(ws.cell(4, c, h))
    ws.row_dimensions[4].height = 40

    PROP_COLS = {4, 5, 6}  
    INT_COLS  = {3, 7, 8}  

    for i, row in df.iterrows():
        alt = i % 2 == 1
        row_vals = [
            row["Company ID"], row["Year"], row["#KAMs"],
            row["%RepKAM_prior_years"], row["%RepKAM_last_year"],
            row["KAM_Similarity"],
            row["#NewKAM_prior_years"], row["#NewKAM_last_year"],
        ]
        for j, v in enumerate(row_vals, 1):
            # None / NaN → tampilkan kosong (tidak terdefinisi untuk tahun pertama)
            if v is None or (isinstance(v, float) and np.isnan(v)):
                cell_val = "N/A"
                fmt      = None
            elif j in PROP_COLS:
                cell_val = v
                fmt      = "0.0000"   
            elif j in INT_COLS:
                cell_val = int(v)
                fmt      = "0"
            else:
                cell_val = v
                fmt      = None
            sc(ws.cell(i + 5, j, cell_val), alt, fmt)

    for i, w in enumerate([25, 8, 8, 22, 20, 26, 22, 20], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"

    ws2 = wb.create_sheet("Summary Statistics")
    ws2.append(["Summary Statistics by Year & Metric"])
    ws2.merge_cells("A1:J1")
    ws2["A1"].font      = Font(name="Arial", bold=True, size=14, color="1F4E79")
    ws2["A1"].alignment = CEN
    ws2.row_dimensions[1].height = 28
    ws2.append([])

    SUM_HEADS = [
        "Metric", "Year", "N", "Mean", "Median",
        "Std Dev", "Min", "Max", "25th %ile", "75th %ile",
    ]
    ws2.append(SUM_HEADS)
    for c, h in enumerate(SUM_HEADS, 1):
        sh(ws2.cell(3, c, h))
    ws2.row_dimensions[3].height = 30

    metric_map = {
        "%RepKAM_prior_years" : True,
        "%RepKAM_last_year"   : True,
        "KAM_Similarity"      : True,
        "#NewKAM_prior_years" : False,
        "#NewKAM_last_year"   : False,
    }
    sr = 4
    for mname, is_pct in metric_map.items():
        for yr in TARGET_YEARS:
            sub = df[df["Year"] == yr][mname].dropna()
            row_data = [
                mname, yr, len(sub),
                sub.mean()           if len(sub) > 0 else "",
                sub.median()         if len(sub) > 0 else "",
                sub.std()            if len(sub) > 1 else "",
                sub.min()            if len(sub) > 0 else "",
                sub.max()            if len(sub) > 0 else "",
                sub.quantile(0.25)   if len(sub) > 0 else "",
                sub.quantile(0.75)   if len(sub) > 0 else "",
            ]
            alt = sr % 2 == 0
            for j, v in enumerate(row_data, 1):
                fmt = "0.0000" if is_pct and j >= 4 and v != "" else None
                sc(ws2.cell(sr, j, v), alt, fmt)
            sr += 1
        ws2.cell(sr, 1).value = ""
        sr += 1

    for i, w in enumerate([25, 8, 6, 10, 10, 10, 10, 10, 10, 10], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A4"

    ws3 = wb.create_sheet("Charts")
    ws3["A1"].value = "Visual Analysis — KAM Similarity"
    ws3["A1"].font  = Font(name="Arial", bold=True, size=15, color="1F4E79")
    ws3.merge_cells("A1:N1")
    ws3["A1"].alignment = CEN
    ws3.row_dimensions[1].height = 30

    row_i = 3
    for label, img_path in chart_paths.items():
        if os.path.exists(img_path):
            img        = XLImage(img_path)
            img.width  = 760
            img.height = 456
            ws3.add_image(img, f"A{row_i}")
            ws3.cell(row_i, 15).value = label
            row_i += 27

    wb.save(output_path)
    print(f">> Full Report tersimpan di: {output_path}")